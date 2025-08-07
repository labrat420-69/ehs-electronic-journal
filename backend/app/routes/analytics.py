"""
Analytics dashboard routes for customizable graphs and data visualization
"""

from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, desc, asc
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.utils import PlotlyJSONEncoder
import json
import io
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.database import get_db
from app.auth.jwt_handler import get_optional_user
from app.models.analytics import GraphPreset, DashboardReminder, DepartmentNote, WasteBox, WasteItem
from app.models.dashboard import DashboardPreferences
from app.models.chemical_inventory import ChemicalInventoryLog, ChemicalInventoryHistory
from app.models.reagents import MMReagents, PbReagents, TCLPReagents
from app.models.standards import MMStandards, FlameAAStandards  
from app.models.equipment import Equipment, PipetteLog, WaterConductivityTests
from app.models.maintenance import ICPOESMaintenanceLog

# Import templates
from pathlib import Path
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
templates = Jinja2Templates(directory=str(PROJECT_ROOT / "frontend" / "templates"))

router = APIRouter()

# Pydantic models for API
class GraphPresetCreate(BaseModel):
    name: str
    description: Optional[str] = None
    graph_type: str
    x_axis_field: str
    y_axis_field: str
    data_source: str
    config: Optional[Dict[str, Any]] = None
    is_public: bool = False

class ReminderCreate(BaseModel):
    title: str
    description: Optional[str] = None
    reminder_type: str
    due_date: datetime
    priority: str = "medium"
    assigned_to: Optional[int] = None

class DepartmentNoteCreate(BaseModel):
    title: str
    content: str
    note_type: str = "general"
    is_pinned: bool = False
    is_public: bool = True
    department: Optional[str] = None

class WasteBoxCreate(BaseModel):
    box_id: str
    coc_job_id: Optional[str] = None
    box_type: str
    size: str
    location: str

class WasteItemCreate(BaseModel):
    item_name: str
    description: Optional[str] = None
    waste_type: str
    quantity: Optional[str] = None
    coc_job_id: Optional[str] = None
    sample_id: Optional[str] = None
    is_extra_sample: bool = False
    waste_box_id: int

class DashboardPreferencesUpdate(BaseModel):
    layout_type: Optional[str] = None
    chart_positions: Optional[List[Dict[str, Any]]] = None
    default_chart_type: Optional[str] = None
    default_data_source: Optional[str] = None
    auto_refresh: Optional[bool] = None
    refresh_interval: Optional[int] = None
    show_sidebar: Optional[bool] = None
    sidebar_collapsed: Optional[bool] = None
    theme_preference: Optional[str] = None
    default_date_range: Optional[str] = None
    max_data_points: Optional[int] = None
    saved_charts: Optional[List[Dict[str, Any]]] = None

@router.get("/analytics", response_class=HTMLResponse)
async def analytics_dashboard(
    request: Request,
    db: Session = Depends(get_db)
):
    """Analytics dashboard page with customizable graphs"""
    current_user = await get_optional_user(request, db)
    
    # Get or create user dashboard preferences
    user_preferences = None
    if current_user:
        user_preferences = db.query(DashboardPreferences).filter(
            DashboardPreferences.user_id == current_user.id
        ).first()
        
        if not user_preferences:
            # Create default preferences for new user
            default_prefs = DashboardPreferences()
            default_data = default_prefs.get_default_preferences()
            
            user_preferences = DashboardPreferences(
                user_id=current_user.id,
                layout_type=default_data["layout_type"],
                chart_positions=default_data["chart_positions"],
                default_chart_type=default_data["default_chart_type"],
                default_data_source=default_data["default_data_source"],
                auto_refresh=default_data["auto_refresh"],
                refresh_interval=default_data["refresh_interval"],
                show_sidebar=default_data["show_sidebar"],
                sidebar_collapsed=default_data["sidebar_collapsed"],
                theme_preference=default_data["theme_preference"],
                default_date_range=default_data["default_date_range"],
                max_data_points=default_data["max_data_points"],
                saved_charts=default_data["saved_charts"]
            )
            db.add(user_preferences)
            db.commit()
            db.refresh(user_preferences)
    
    # Get user's graph presets
    user_presets = []
    public_presets = []
    if current_user:
        user_presets = db.query(GraphPreset).filter(
            GraphPreset.created_by == current_user.id
        ).all()
        public_presets = db.query(GraphPreset).filter(
            and_(GraphPreset.is_public == True, GraphPreset.created_by != current_user.id)
        ).limit(10).all()
    else:
        public_presets = db.query(GraphPreset).filter(
            GraphPreset.is_public == True
        ).limit(10).all()
    
    # Get available data sources and fields
    data_sources = get_available_data_sources()
    
    # Get recent reminders
    recent_reminders = []
    if current_user:
        recent_reminders = db.query(DashboardReminder).filter(
            or_(
                DashboardReminder.created_by == current_user.id,
                DashboardReminder.assigned_to == current_user.id
            ),
            DashboardReminder.status == "active"
        ).order_by(DashboardReminder.due_date).limit(5).all()
    
    # Get recent notes
    recent_notes = db.query(DepartmentNote).filter(
        DepartmentNote.is_public == True
    ).order_by(desc(DepartmentNote.created_at)).limit(5).all()
    
    context = {
        "request": request,
        "title": "Analytics Dashboard - EHS Electronic Journal",
        "current_user": current_user,
        "user_presets": [preset.to_dict() for preset in user_presets],
        "public_presets": [preset.to_dict() for preset in public_presets],
        "data_sources": data_sources,
        "recent_reminders": [reminder.to_dict() for reminder in recent_reminders],
        "recent_notes": [note.to_dict() for note in recent_notes],
        "user_preferences": user_preferences.to_dict() if user_preferences else None
    }
    
    return templates.TemplateResponse("analytics/dashboard.html", context)

@router.get("/api/analytics/data-sources")
async def get_data_sources():
    """Get available data sources and their fields"""
    return get_available_data_sources()

@router.get("/api/analytics/graph-data")
async def get_graph_data(
    data_source: str,
    x_field: str,
    y_field: str,
    graph_type: str = "line",
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get data for graph based on source and field selections"""
    try:
        data = await fetch_graph_data(db, data_source, x_field, y_field, limit)
        
        # Create the graph based on type
        graph_html = create_graph(data, x_field, y_field, graph_type, data_source)
        
        return {
            "success": True,
            "graph_html": graph_html,
            "data_points": len(data),
            "x_field": x_field,
            "y_field": y_field,
            "data_source": data_source
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/api/analytics/presets")
async def create_graph_preset(
    preset: GraphPresetCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    """Create a new graph preset"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    db_preset = GraphPreset(
        name=preset.name,
        description=preset.description,
        graph_type=preset.graph_type,
        x_axis_field=preset.x_axis_field,
        y_axis_field=preset.y_axis_field,
        data_source=preset.data_source,
        config=preset.config,
        is_public=preset.is_public,
        created_by=current_user.id
    )
    
    db.add(db_preset)
    db.commit()
    db.refresh(db_preset)
    
    return {"success": True, "preset": db_preset.to_dict()}

@router.get("/api/analytics/presets")
async def get_graph_presets(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get user's graph presets"""
    current_user = await get_optional_user(request, db)
    
    if current_user:
        user_presets = db.query(GraphPreset).filter(
            GraphPreset.created_by == current_user.id
        ).all()
        public_presets = db.query(GraphPreset).filter(
            and_(GraphPreset.is_public == True, GraphPreset.created_by != current_user.id)
        ).all()
    else:
        user_presets = []
        public_presets = db.query(GraphPreset).filter(
            GraphPreset.is_public == True
        ).all()
    
    return {
        "user_presets": [preset.to_dict() for preset in user_presets],
        "public_presets": [preset.to_dict() for preset in public_presets]
    }

@router.delete("/api/analytics/presets/{preset_id}")
async def delete_graph_preset(
    preset_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Delete a graph preset"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    preset = db.query(GraphPreset).filter(
        and_(GraphPreset.id == preset_id, GraphPreset.created_by == current_user.id)
    ).first()
    
    if not preset:
        raise HTTPException(status_code=404, detail="Preset not found")
    
    db.delete(preset)
    db.commit()
    
    return {"success": True}

@router.get("/api/analytics/export-excel")
async def export_graph_excel(
    data_source: str,
    x_field: str,
    y_field: str,
    filename: str = "analytics_data",
    db: Session = Depends(get_db)
):
    """Export graph data to Excel with precise formatting"""
    try:
        data = await fetch_graph_data(db, data_source, x_field, y_field, limit=10000)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel writer with styling
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Analytics Data', index=False)
            
            # Get the worksheet to apply styling
            worksheet = writer.sheets['Analytics Data']
            
            # Style the header
            header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Return as streaming response
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}.xlsx"'
        }
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/api/analytics/template/{data_source}")
async def download_import_template(data_source: str):
    """Download Excel template for CSV import"""
    try:
        template_data = get_import_template_data(data_source)
        
        # Create Excel file in memory
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"{data_source.title()} Import Template"
        
        # Add headers
        headers = template_data["headers"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='f0f8ff', end_color='f0f8ff', fill_type='solid')
        
        # Add sample data
        sample_data = template_data["sample_data"]
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Add instructions sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = template_data["instructions"]
        for row_idx, instruction in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_idx, column=1, value=instruction)
        
        # Auto-adjust column widths
        for worksheet in workbook.worksheets:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="{data_source}_import_template.xlsx"'
        }
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Helper functions

def get_available_data_sources():
    """Get available data sources and their fields for graph creation"""
    return {
        "chemical_inventory": {
            "name": "Chemical Inventory",
            "fields": {
                "created_at": "Date Created",
                "updated_at": "Date Updated", 
                "current_quantity": "Current Quantity",
                "expiration_date": "Expiration Date",
                "received_date": "Received Date",
                "chemical_name": "Chemical Name",
                "manufacturer": "Manufacturer",
                "storage_location": "Storage Location",
                "hazard_class": "Hazard Class",
                "is_hazardous": "Is Hazardous"
            }
        },
        "chemical_history": {
            "name": "Chemical History",
            "fields": {
                "changed_at": "Change Date",
                "quantity_change": "Quantity Change",
                "remaining_quantity": "Remaining Quantity",
                "action": "Action",
                "chemical_id": "Chemical ID"
            }
        },
        "reagents": {
            "name": "Reagents",
            "fields": {
                "preparation_date": "Preparation Date",
                "expiry_date": "Expiry Date",
                "volume_prepared": "Volume Prepared",
                "reagent_type": "Reagent Type",
                "is_active": "Is Active"
            }
        },
        "standards": {
            "name": "Standards",
            "fields": {
                "preparation_date": "Preparation Date",
                "expiry_date": "Expiry Date",
                "concentration": "Concentration",
                "volume": "Volume",
                "standard_type": "Standard Type"
            }
        },
        "equipment": {
            "name": "Equipment",
            "fields": {
                "calibration_date": "Calibration Date",
                "next_calibration_due": "Next Calibration Due",
                "equipment_name": "Equipment Name",
                "model": "Model",
                "serial_number": "Serial Number"
            }
        },
        "pipette_tests": {
            "name": "Pipette Tests", 
            "fields": {
                "test_date": "Test Date",
                "accuracy_result": "Accuracy Result",
                "precision_result": "Precision Result",
                "pipette_id": "Pipette ID",
                "volume_setting": "Volume Setting",
                "calibration_status": "Calibration Status"
            }
        },
        "water_tests": {
            "name": "Water Conductivity Tests",
            "fields": {
                "test_date": "Test Date",
                "conductivity_reading": "Conductivity Reading",
                "result_status": "Result Status",
                "sample_source": "Sample Source",
                "temperature": "Temperature"
            }
        },
        "maintenance": {
            "name": "Maintenance Logs",
            "fields": {
                "maintenance_date": "Maintenance Date",
                "maintenance_category": "Category",
                "hours_spent": "Hours Spent",
                "equipment_status": "Equipment Status"
            }
        },
        "waste_boxes": {
            "name": "Waste Boxes",
            "fields": {
                "created_date": "Created Date",
                "filled_date": "Filled Date",
                "disposed_date": "Disposed Date",
                "fill_percentage": "Fill Percentage",
                "box_type": "Box Type",
                "size": "Size",
                "status": "Status"
            }
        }
    }

async def fetch_graph_data(db: Session, data_source: str, x_field: str, y_field: str, limit: int = 100):
    """Fetch data from database based on source and field selections"""
    
    if data_source == "chemical_inventory":
        query = db.query(ChemicalInventoryLog).filter(ChemicalInventoryLog.is_active == True)
        records = query.limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                # Convert datetime to string for JSON serialization
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    elif data_source == "chemical_history":
        query = db.query(ChemicalInventoryHistory)
        records = query.order_by(desc(ChemicalInventoryHistory.changed_at)).limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    elif data_source == "pipette_tests":
        query = db.query(PipetteLog)
        records = query.order_by(desc(PipetteLog.test_date)).limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    elif data_source == "water_tests":
        query = db.query(WaterConductivityTests)
        records = query.order_by(desc(WaterConductivityTests.test_date)).limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    elif data_source == "maintenance":
        query = db.query(ICPOESMaintenanceLog)
        records = query.order_by(desc(ICPOESMaintenanceLog.maintenance_date)).limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    elif data_source == "waste_boxes":
        query = db.query(WasteBox)
        records = query.order_by(desc(WasteBox.created_date)).limit(limit).all()
        data = []
        for record in records:
            x_val = getattr(record, x_field, None)
            y_val = getattr(record, y_field, None)
            if x_val is not None and y_val is not None:
                if isinstance(x_val, datetime):
                    x_val = x_val.isoformat()
                if isinstance(y_val, datetime):
                    y_val = y_val.isoformat()
                data.append({x_field: x_val, y_field: y_val})
    
    else:
        raise ValueError(f"Unsupported data source: {data_source}")
    
    return data

def create_graph(data, x_field, y_field, graph_type, data_source):
    """Create Plotly graph HTML"""
    if not data:
        return "<div>No data available for selected fields</div>"
    
    # Convert data to DataFrame for easier plotting
    df = pd.DataFrame(data)
    
    # Create figure based on graph type
    if graph_type == "line":
        fig = px.line(df, x=x_field, y=y_field, title=f"{data_source.title()}: {y_field} vs {x_field}")
        # Add red/green trend colors
        fig.update_traces(line=dict(color='#1a73e8', width=2))
    
    elif graph_type == "bar":
        fig = px.bar(df, x=x_field, y=y_field, title=f"{data_source.title()}: {y_field} vs {x_field}")
        fig.update_traces(marker_color='#4285f4')
    
    elif graph_type == "area":
        fig = px.area(df, x=x_field, y=y_field, title=f"{data_source.title()}: {y_field} vs {x_field}")
        fig.update_traces(fill='tonexty', fillcolor='rgba(26,115,232,0.3)')
    
    elif graph_type == "scatter":
        fig = px.scatter(df, x=x_field, y=y_field, title=f"{data_source.title()}: {y_field} vs {x_field}")
        fig.update_traces(marker=dict(color='#ea4335', size=8))
    
    else:  # Default to line
        fig = px.line(df, x=x_field, y=y_field, title=f"{data_source.title()}: {y_field} vs {x_field}")
    
    # Apply crypto-style theme
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white', family='Arial, sans-serif'),
        title=dict(font=dict(size=16, color='#1a73e8')),
        xaxis=dict(gridcolor='rgba(255,255,255,0.1)', showgrid=True),
        yaxis=dict(gridcolor='rgba(255,255,255,0.1)', showgrid=True),
        height=400
    )
    
    # Convert to HTML
    graph_html = fig.to_html(include_plotlyjs='cdn', div_id=f"graph-{data_source}")
    
    return graph_html

def get_import_template_data(data_source: str):
    """Get template data for CSV import"""
    templates = {
        "chemical_inventory": {
            "headers": [
                "chemical_name", "cas_number", "manufacturer", "catalog_number", 
                "lot_number", "container_size", "current_quantity", "unit",
                "storage_location", "storage_temperature", "hazard_class",
                "received_date", "expiration_date", "is_hazardous", "safety_notes"
            ],
            "sample_data": [
                ["Hydrochloric Acid", "7647-01-0", "Fisher Scientific", "A144-500", 
                 "ABC123", "500ml", "450", "ml", "Acid Cabinet A1", "Room Temperature",
                 "Corrosive", "2024-01-15", "2026-01-15", "TRUE", "Corrosive to skin and eyes"],
                ["Sodium Chloride", "7647-14-5", "Sigma-Aldrich", "S9888-500G",
                 "XYZ789", "500g", "350", "g", "Salt Storage", "Room Temperature",
                 "None", "2024-02-01", "2027-02-01", "FALSE", "Non-hazardous reagent grade"]
            ],
            "instructions": [
                "Chemical Inventory Import Template",
                "",
                "Instructions:",
                "1. Fill in all required fields (chemical_name, current_quantity, unit are mandatory)",
                "2. Use TRUE/FALSE for boolean fields (is_hazardous)",
                "3. Date format: YYYY-MM-DD (e.g., 2024-12-31)",
                "4. Quantity should be numeric",
                "5. Save as CSV file before uploading",
                "",
                "Required fields: chemical_name, current_quantity, unit",
                "Optional fields: All others"
            ]
        },
        "waste_boxes": {
            "headers": [
                "box_id", "coc_job_id", "box_type", "size", "location",
                "status", "fill_percentage"
            ],
            "sample_data": [
                ["WB-2024-001", "COC-12345", "hazardous", "medium", "Waste Storage Room A",
                 "active", "25.5"],
                ["WB-2024-002", "COC-12346", "non-hazardous", "large", "Waste Storage Room B",
                 "full", "100.0"]
            ],
            "instructions": [
                "Waste Box Import Template",
                "",
                "Instructions:",
                "1. box_id must be unique",
                "2. box_type: hazardous, non-hazardous, glass, sharps",
                "3. size: small, medium, large",
                "4. status: active, full, disposed, in_storage",
                "5. fill_percentage: 0.0 to 100.0",
                "",
                "Required fields: box_id, box_type, size, location"
            ]
        }
    }
    
    return templates.get(data_source, {
        "headers": ["field1", "field2"],
        "sample_data": [["value1", "value2"]],
        "instructions": ["Template not available for this data source"]
    })

# Additional routes for reminders, notes, and waste management will be added in separate route files

# Dashboard preferences routes
@router.get("/api/analytics/preferences")
async def get_dashboard_preferences(
    request: Request,
    db: Session = Depends(get_db)
):
    """Get user dashboard preferences"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    preferences = db.query(DashboardPreferences).filter(
        DashboardPreferences.user_id == current_user.id
    ).first()
    
    if not preferences:
        # Return default preferences
        default_prefs = DashboardPreferences()
        return {"success": True, "preferences": default_prefs.get_default_preferences()}
    
    return {"success": True, "preferences": preferences.to_dict()}

@router.post("/api/analytics/preferences")
async def update_dashboard_preferences(
    preferences_update: DashboardPreferencesUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """Update user dashboard preferences"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get or create preferences
    preferences = db.query(DashboardPreferences).filter(
        DashboardPreferences.user_id == current_user.id
    ).first()
    
    if not preferences:
        preferences = DashboardPreferences(user_id=current_user.id)
        db.add(preferences)
    
    # Update fields that were provided
    update_data = preferences_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(preferences, field, value)
    
    db.commit()
    db.refresh(preferences)
    
    return {"success": True, "preferences": preferences.to_dict()}

@router.post("/api/analytics/preferences/save-chart")
async def save_chart_configuration(
    chart_config: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db)
):
    """Save a chart configuration to user preferences"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    preferences = db.query(DashboardPreferences).filter(
        DashboardPreferences.user_id == current_user.id
    ).first()
    
    if not preferences:
        preferences = DashboardPreferences(user_id=current_user.id)
        db.add(preferences)
    
    # Add chart to saved charts
    saved_charts = preferences.saved_charts or []
    
    # Add timestamp and unique ID
    import uuid
    chart_config["id"] = str(uuid.uuid4())
    chart_config["created_at"] = datetime.utcnow().isoformat()
    
    saved_charts.append(chart_config)
    preferences.saved_charts = saved_charts
    
    db.commit()
    db.refresh(preferences)
    
    return {"success": True, "chart_id": chart_config["id"]}

@router.delete("/api/analytics/preferences/saved-charts/{chart_id}")
async def delete_saved_chart(
    chart_id: str,
    request: Request,
    db: Session = Depends(get_db)
):
    """Delete a saved chart configuration"""
    current_user = await get_optional_user(request, db)
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    preferences = db.query(DashboardPreferences).filter(
        DashboardPreferences.user_id == current_user.id
    ).first()
    
    if not preferences:
        raise HTTPException(status_code=404, detail="Preferences not found")
    
    saved_charts = preferences.saved_charts or []
    original_length = len(saved_charts)
    
    # Remove chart with matching ID
    saved_charts = [chart for chart in saved_charts if chart.get("id") != chart_id]
    
    if len(saved_charts) == original_length:
        raise HTTPException(status_code=404, detail="Chart not found")
    
    preferences.saved_charts = saved_charts
    db.commit()
    
    return {"success": True}