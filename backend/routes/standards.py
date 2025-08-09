"""
Standards routes - MM, FlameAA, and Mercury
"""

from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, validator
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from backend.database import get_db
from backend.models.standards import (
    MMStandards, MMStandardsHistory,
    FlameAAStandards, FlameAAStandardsHistory
)
# Import Mercury standards from reagents model
from backend.models.reagents import MercuryStandards, MercuryStandardsHistory
from backend.models.user import User
from backend.auth.jwt_handler import get_current_user, require_permissions

# Import templates - use the same pattern as main.py
from fastapi.templating import Jinja2Templates

from backend.utils.template_helpers import template_functions

templates = Jinja2Templates(directory="frontend/templates")
# Add template helper functions for robust role-based access control
templates.env.globals.update(template_functions)

router = APIRouter(prefix="/standards", tags=["Standards"])

# Pydantic models for MM Standards
class MMStandardCreate(BaseModel):
    standard_name: str
    batch_number: str
    standard_type: str  # QC, Calibration, Spike, etc.
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    target_concentration: float
    actual_concentration: Optional[float] = None
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    total_volume: float
    elements: Optional[str] = None  # JSON string
    verification_method: Optional[str] = None
    certified: bool = False
    certificate_number: Optional[str] = None
    notes: Optional[str] = None

    @validator('total_volume', 'target_concentration')
    def values_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Value must be positive')
        return v

class MMStandardUpdate(BaseModel):
    standard_name: Optional[str] = None
    standard_type: Optional[str] = None
    expiration_date: Optional[datetime] = None
    actual_concentration: Optional[float] = None
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    elements: Optional[str] = None
    verification_method: Optional[str] = None
    certified: Optional[bool] = None
    certificate_number: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

# Pydantic models for FlameAA Standards
class FlameAAStandardCreate(BaseModel):
    standard_name: str
    batch_number: str
    element: str  # Ca, Mg, Na, K, etc.
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    concentration: float
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    total_volume: float
    verified_concentration: Optional[float] = None
    verification_method: Optional[str] = None
    notes: Optional[str] = None

    @validator('total_volume', 'concentration')
    def values_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Value must be positive')
        return v

class FlameAAStandardUpdate(BaseModel):
    standard_name: Optional[str] = None
    element: Optional[str] = None
    expiration_date: Optional[datetime] = None
    concentration: Optional[float] = None
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    verified_concentration: Optional[float] = None
    verification_method: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

class VolumeUpdate(BaseModel):
    volume_change: float
    reason: str
    notes: Optional[str] = None

    @validator('volume_change')
    def volume_change_not_zero(cls, v):
        if v == 0:
            raise ValueError('Volume change cannot be zero')
        return v

# Pydantic models for Mercury Standards
class MercuryStandardCreate(BaseModel):
    standard_name: str
    batch_number: str
    standard_type: str  # QC, Calibration, Spike, etc.
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    target_concentration: float
    actual_concentration: Optional[float] = None
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    total_volume: float
    elements: Optional[str] = None  # JSON string
    verification_method: Optional[str] = None
    certified: bool = False
    certificate_number: Optional[str] = None
    notes: Optional[str] = None

    @validator('total_volume', 'target_concentration')
    def volume_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Value must be positive')
        return v

class MercuryStandardUpdate(BaseModel):
    standard_name: Optional[str] = None
    standard_type: Optional[str] = None
    expiration_date: Optional[datetime] = None
    target_concentration: Optional[float] = None
    actual_concentration: Optional[float] = None
    matrix: Optional[str] = None
    source_material: Optional[str] = None
    dilution_factor: Optional[float] = None
    elements: Optional[str] = None
    verification_method: Optional[str] = None
    certified: Optional[bool] = None
    certificate_number: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

# MM Standards Routes
@router.get("/mm", response_class=HTMLResponse)
async def mm_standards_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """MM Standards list page"""
    standards = db.query(MMStandards).filter(
        MMStandards.is_active == True
    ).order_by(MMStandards.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "MM Standards - EHS Electronic Journal",
        "standards": standards,
        "current_user": current_user,
        "standard_type": "mm",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("standards/list.html", context)

@router.get("/mm/add", response_class=HTMLResponse)
async def add_mm_standard_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add MM standard form"""
    context = {
        "request": request,
        "title": "Add MM Standard - EHS Electronic Journal",
        "current_user": current_user,
        "standard_type": "mm"
    }
    
    return templates.TemplateResponse("standards/add.html", context)

@router.post("/mm/api/", response_model=dict)
async def create_mm_standard(
    standard: MMStandardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new MM standard"""
    
    # Check if batch number already exists
    existing = db.query(MMStandards).filter(MMStandards.batch_number == standard.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        # Set initial volume and current volume to the same value
        standard_data = standard.dict()
        standard_data['initial_volume'] = standard_data['total_volume']
        standard_data['current_volume'] = standard_data['total_volume']
        
        db_standard = MMStandards(**standard_data, prepared_by=current_user.id)
        db.add(db_standard)
        db.commit()
        db.refresh(db_standard)
        
        # Create history entry
        history_entry = MMStandardsHistory(
            standard_id=db_standard.id,
            action="created",
            new_value=f"MM Standard {db_standard.standard_name} prepared",
            notes="Initial standard preparation",
            remaining_volume=db_standard.current_volume,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "MM Standard created successfully",
            "standard": db_standard.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating MM standard: {str(e)}"
        )

@router.get("/mm/api/", response_model=List[dict])
async def list_mm_standards(
    active_only: bool = True,
    standard_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """List all MM standards"""
    
    query = db.query(MMStandards)
    if active_only:
        query = query.filter(MMStandards.is_active == True)
    if standard_type:
        query = query.filter(MMStandards.standard_type == standard_type)
    
    standards = query.order_by(MMStandards.preparation_date.desc()).all()
    return [standard.to_dict() for standard in standards]

@router.get("/mm/{standard_id}", response_class=HTMLResponse)
async def mm_standard_detail(
    standard_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """MM standard detail page"""
    
    standard = db.query(MMStandards).filter(MMStandards.id == standard_id).first()
    if not standard:
        raise HTTPException(status_code=404, detail="Standard not found")
    
    history = db.query(MMStandardsHistory).filter(
        MMStandardsHistory.standard_id == standard_id
    ).order_by(MMStandardsHistory.changed_at.desc()).all()
    
    context = {
        "request": request,
        "title": f"{standard.standard_name} - MM Standard Details",
        "standard": standard,
        "history": history,
        "current_user": current_user,
        "standard_type": "mm"
    }
    
    return templates.TemplateResponse("standards/detail.html", context)

@router.put("/mm/api/{standard_id}", response_model=dict)
async def update_mm_standard(
    standard_id: int,
    standard_update: MMStandardUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["update"]))
):
    """Update MM standard"""
    
    db_standard = db.query(MMStandards).filter(MMStandards.id == standard_id).first()
    if not db_standard:
        raise HTTPException(status_code=404, detail="Standard not found")
    
    # Track changes for history
    changes = []
    update_data = standard_update.dict(exclude_unset=True)
    
    for field, new_value in update_data.items():
        old_value = getattr(db_standard, field)
        if old_value != new_value:
            changes.append({
                "field": field,
                "old_value": str(old_value) if old_value else None,
                "new_value": str(new_value) if new_value else None
            })
            setattr(db_standard, field, new_value)
    
    if not changes:
        return {
            "success": True,
            "message": "No changes detected",
            "standard": db_standard.to_dict()
        }
    
    try:
        db.commit()
        
        # Create history entries for each change
        for change in changes:
            history_entry = MMStandardsHistory(
                standard_id=standard_id,
                action="updated",
                field_changed=change["field"],
                old_value=change["old_value"],
                new_value=change["new_value"],
                remaining_volume=db_standard.current_volume,
                changed_by=current_user.id
            )
            db.add(history_entry)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"MM Standard updated successfully. {len(changes)} field(s) modified.",
            "standard": db_standard.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error updating standard: {str(e)}"
        )

@router.patch("/mm/api/{standard_id}/volume", response_model=dict)
async def update_mm_volume(
    standard_id: int,
    volume_update: VolumeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["update"]))
):
    """Update MM standard volume"""
    
    db_standard = db.query(MMStandards).filter(MMStandards.id == standard_id).first()
    if not db_standard:
        raise HTTPException(status_code=404, detail="Standard not found")
    
    old_volume = db_standard.current_volume
    new_volume = old_volume + volume_update.volume_change
    
    if new_volume < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Insufficient volume available"
        )
    
    try:
        db_standard.current_volume = new_volume
        db.commit()
        
        # Create history entry for volume change
        action = "volume_added" if volume_update.volume_change > 0 else "volume_used"
        history_entry = MMStandardsHistory(
            standard_id=standard_id,
            action=action,
            field_changed="current_volume",
            old_value=str(old_volume),
            new_value=str(new_volume),
            volume_used=volume_update.volume_change,
            remaining_volume=new_volume,
            reason=volume_update.reason,
            notes=volume_update.notes,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": f"Volume updated: {volume_update.volume_change:+.3f} mL",
            "standard": db_standard.to_dict(),
            "old_volume": float(old_volume),
            "new_volume": float(new_volume)
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error updating volume: {str(e)}"
        )

# FlameAA Standards Routes
@router.get("/flameaa", response_class=HTMLResponse)
async def flameaa_standards_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """FlameAA Standards list page"""
    standards = db.query(FlameAAStandards).filter(
        FlameAAStandards.is_active == True
    ).order_by(FlameAAStandards.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "FlameAA Standards - EHS Electronic Journal",
        "standards": standards,
        "current_user": current_user,
        "standard_type": "flameaa",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("standards/list.html", context)

@router.get("/flameaa/add", response_class=HTMLResponse)
async def add_flameaa_standard_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add FlameAA standard form"""
    context = {
        "request": request,
        "title": "Add FlameAA Standard - EHS Electronic Journal",
        "current_user": current_user,
        "standard_type": "flameaa"
    }
    
    return templates.TemplateResponse("standards/add.html", context)

@router.post("/flameaa/api/", response_model=dict)
async def create_flameaa_standard(
    standard: FlameAAStandardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new FlameAA standard"""
    
    # Check if batch number already exists
    existing = db.query(FlameAAStandards).filter(FlameAAStandards.batch_number == standard.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        # Set initial volume and current volume to the same value
        standard_data = standard.dict()
        standard_data['initial_volume'] = standard_data['total_volume']
        standard_data['current_volume'] = standard_data['total_volume']
        
        db_standard = FlameAAStandards(**standard_data, prepared_by=current_user.id)
        db.add(db_standard)
        db.commit()
        db.refresh(db_standard)
        
        # Create history entry
        history_entry = FlameAAStandardsHistory(
            standard_id=db_standard.id,
            action="created",
            new_value=f"FlameAA Standard {db_standard.standard_name} prepared",
            notes="Initial standard preparation",
            remaining_volume=db_standard.current_volume,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "FlameAA Standard created successfully",
            "standard": db_standard.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating FlameAA standard: {str(e)}"
        )

# Generic routes for all standard types
@router.get("/api/", response_model=dict)
async def list_all_standards(
    standard_type: Optional[str] = None,
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """List all standards across types"""
    
    result = {
        "mm_standards": [],
        "flameaa_standards": [],
        "mercury_standards": []
    }
    
    if not standard_type or standard_type.lower() == "mm":
        query = db.query(MMStandards)
        if active_only:
            query = query.filter(MMStandards.is_active == True)
        result["mm_standards"] = [s.to_dict() for s in query.all()]
    
    if not standard_type or standard_type.lower() == "flameaa":
        query = db.query(FlameAAStandards)
        if active_only:
            query = query.filter(FlameAAStandards.is_active == True)
        result["flameaa_standards"] = [s.to_dict() for s in query.all()]
    
    if not standard_type or standard_type.lower() == "mercury":
        query = db.query(MercuryStandards)
        if active_only:
            query = query.filter(MercuryStandards.is_active == True)
        result["mercury_standards"] = [s.to_dict() for s in query.all()]
    
    return result

# Mercury Standards Routes
@router.get("/mercury", response_class=HTMLResponse)
async def mercury_standards_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Mercury Standards list page"""
    standards = db.query(MercuryStandards).filter(
        MercuryStandards.is_active == True
    ).order_by(MercuryStandards.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "Mercury Standards - EHS Electronic Journal",
        "standards": standards,
        "current_user": current_user,
        "standard_type": "mercury",
        "reagent_type": "Mercury",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("standards/list.html", context)

@router.get("/mercury/add", response_class=HTMLResponse)
async def add_mercury_standard_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add Mercury standard form"""
    context = {
        "request": request,
        "title": "Add Mercury Standard - EHS Electronic Journal",
        "current_user": current_user,
        "standard_type": "mercury",
        "reagent_type": "Mercury"
    }
    
    return templates.TemplateResponse("standards/add.html", context)

@router.post("/mercury/api/", response_model=dict)
async def create_mercury_standard(
    standard: MercuryStandardCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new Mercury standard"""
    
    # Check if batch number already exists
    existing = db.query(MercuryStandards).filter(MercuryStandards.batch_number == standard.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        db_standard = MercuryStandards(**standard.dict(), prepared_by=current_user.id)
        db.add(db_standard)
        db.commit()
        db.refresh(db_standard)
        
        # Create history entry
        history_entry = MercuryStandardsHistory(
            standard_id=db_standard.id,
            action="created",
            new_value=f"Mercury Standard {db_standard.standard_name} prepared",
            notes="Initial standard preparation",
            changed_by=current_user.id
        )
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "Mercury Standard created successfully",
            "standard": db_standard.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating Mercury standard: {str(e)}"
        )

# Export endpoints for all standard types
@router.get("/mm/export")
async def export_mm_standards(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export MM standards to Excel"""
    standards = db.query(MMStandards).filter(MMStandards.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for standard in standards:
        data.append({
            'ID': standard.id,
            'Standard Name': standard.standard_name,
            'Batch Number': standard.batch_number,
            'Standard Type': standard.standard_type,
            'Preparation Date': standard.preparation_date.strftime('%Y-%m-%d') if standard.preparation_date else '',
            'Expiration Date': standard.expiration_date.strftime('%Y-%m-%d') if standard.expiration_date else '',
            'Target Concentration': float(standard.target_concentration) if standard.target_concentration else 0,
            'Actual Concentration': float(standard.actual_concentration) if standard.actual_concentration else '',
            'Total Volume (mL)': float(standard.total_volume) if standard.total_volume else 0,
            'Matrix': standard.matrix or '',
            'Source Material': standard.source_material or '',
            'Dilution Factor': float(standard.dilution_factor) if standard.dilution_factor else '',
            'Elements': standard.elements or '',
            'Certified': standard.certified,
            'Certificate Number': standard.certificate_number or '',
            'Prepared By': standard.preparer.full_name if standard.preparer else '',
            'Notes': standard.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='MM Standards', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['MM Standards']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mm_standards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@router.get("/flameaa/export")
async def export_flameaa_standards(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export FlameAA standards to Excel"""
    standards = db.query(FlameAAStandards).filter(FlameAAStandards.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for standard in standards:
        data.append({
            'ID': standard.id,
            'Standard Name': standard.standard_name,
            'Batch Number': standard.batch_number,
            'Standard Type': standard.standard_type,
            'Preparation Date': standard.preparation_date.strftime('%Y-%m-%d') if standard.preparation_date else '',
            'Expiration Date': standard.expiration_date.strftime('%Y-%m-%d') if standard.expiration_date else '',
            'Target Concentration': float(standard.target_concentration) if standard.target_concentration else 0,
            'Actual Concentration': float(standard.actual_concentration) if standard.actual_concentration else '',
            'Total Volume (mL)': float(standard.total_volume) if standard.total_volume else 0,
            'Matrix': standard.matrix or '',
            'Source Material': standard.source_material or '',
            'Dilution Factor': float(standard.dilution_factor) if standard.dilution_factor else '',
            'Elements': standard.elements or '',
            'Flame Type': standard.flame_type or '',
            'Wavelength': float(standard.wavelength) if standard.wavelength else '',
            'Certified': standard.certified,
            'Certificate Number': standard.certificate_number or '',
            'Prepared By': standard.preparer.full_name if standard.preparer else '',
            'Notes': standard.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='FlameAA Standards', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['FlameAA Standards']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=flameaa_standards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@router.get("/mercury/export")
async def export_mercury_standards(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export Mercury standards to Excel"""
    standards = db.query(MercuryStandards).filter(MercuryStandards.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for standard in standards:
        data.append({
            'ID': standard.id,
            'Standard Name': standard.standard_name,
            'Batch Number': standard.batch_number,
            'Standard Type': standard.standard_type,
            'Preparation Date': standard.preparation_date.strftime('%Y-%m-%d') if standard.preparation_date else '',
            'Expiration Date': standard.expiration_date.strftime('%Y-%m-%d') if standard.expiration_date else '',
            'Target Concentration': float(standard.target_concentration) if standard.target_concentration else 0,
            'Actual Concentration': float(standard.actual_concentration) if standard.actual_concentration else '',
            'Total Volume (mL)': float(standard.total_volume) if standard.total_volume else 0,
            'Matrix': standard.matrix or '',
            'Source Material': standard.source_material or '',
            'Dilution Factor': float(standard.dilution_factor) if standard.dilution_factor else '',
            'Elements': standard.elements or '',
            'Verification Method': standard.verification_method or '',
            'Certified': standard.certified,
            'Certificate Number': standard.certificate_number or '',
            'Prepared By': standard.preparer.full_name if standard.preparer else '',
            'Notes': standard.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Mercury Standards', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['Mercury Standards']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mercury_standards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )