"""
Reagents routes - MM, Pb, TCLP, Mercury
"""

from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, validator
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from backend.database import get_db
from backend.models.reagents import (
    MMReagents, MMReagentsHistory,
    PbReagents, PbReagentsHistory,
    TCLPReagents, TCLPReagentsHistory,
    MercuryReagents, MercuryReagentsHistory
)
from backend.models.user import User
from backend.auth.jwt_handler import get_current_user, require_permissions

# Import templates - use the same pattern as main.py
from fastapi.templating import Jinja2Templates

from backend.utils.template_helpers import template_functions

templates = Jinja2Templates(directory="frontend/templates")
# Add template helper functions for robust role-based access control
templates.env.globals.update(template_functions)

router = APIRouter(prefix="/reagents", tags=["Reagents"])

# Pydantic models for MM Reagents
class MMReagentCreate(BaseModel):
    reagent_name: str
    batch_number: str
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    total_volume: float
    concentration: Optional[str] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    ph_value: Optional[float] = None
    conductivity: Optional[float] = None
    notes: Optional[str] = None

    @validator('total_volume')
    def volume_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Volume must be positive')
        return v

class MMReagentUpdate(BaseModel):
    reagent_name: Optional[str] = None
    expiration_date: Optional[datetime] = None
    concentration: Optional[str] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    ph_value: Optional[float] = None
    conductivity: Optional[float] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

# Pydantic models for Pb Reagents
class PbReagentCreate(BaseModel):
    reagent_name: str
    batch_number: str
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    total_volume: float
    lead_concentration: Optional[float] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    notes: Optional[str] = None

    @validator('total_volume')
    def volume_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Volume must be positive')
        return v

class PbReagentUpdate(BaseModel):
    reagent_name: Optional[str] = None
    expiration_date: Optional[datetime] = None
    lead_concentration: Optional[float] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

# Pydantic models for TCLP Reagents
class TCLPReagentCreate(BaseModel):
    reagent_name: str
    batch_number: str
    reagent_type: str  # Extraction Fluid 1, Extraction Fluid 2, etc.
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    total_volume: float
    ph_target: Optional[float] = None
    final_ph: Optional[float] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    conductivity: Optional[float] = None
    verification_passed: bool = False
    notes: Optional[str] = None

    @validator('total_volume')
    def volume_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Volume must be positive')
        return v

class TCLPReagentUpdate(BaseModel):
    reagent_name: Optional[str] = None
    reagent_type: Optional[str] = None
    expiration_date: Optional[datetime] = None
    final_ph: Optional[float] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    conductivity: Optional[float] = None
    verification_passed: Optional[bool] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

class VolumeUpdate(BaseModel):
    volume_change: float
    reason: str
    notes: Optional[str] = None

    @validator('volume_change')
    def volume_change_not_zero(cls, v):
        if v == 0:
            raise ValueError('Volume change cannot be zero')
        return v

# Pydantic models for Mercury Reagents
class MercuryReagentCreate(BaseModel):
    reagent_name: str
    batch_number: str
    preparation_date: datetime
    expiration_date: Optional[datetime] = None
    total_volume: float
    concentration: Optional[str] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    ph_value: Optional[float] = None
    conductivity: Optional[float] = None
    notes: Optional[str] = None

    @validator('total_volume')
    def volume_must_be_positive(cls, v):
        if v <= 0:
            raise ValueError('Volume must be positive')
        return v

class MercuryReagentUpdate(BaseModel):
    reagent_name: Optional[str] = None
    expiration_date: Optional[datetime] = None
    concentration: Optional[str] = None
    preparation_method: Optional[str] = None
    chemicals_used: Optional[str] = None
    ph_value: Optional[float] = None
    conductivity: Optional[float] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = None

# MM Reagents Routes
@router.get("/mm", response_class=HTMLResponse)
async def mm_reagents_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """MM Reagents list page"""
    reagents = db.query(MMReagents).filter(
        MMReagents.is_active == True
    ).order_by(MMReagents.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "MM Reagents - EHS Electronic Journal",
        "reagents": reagents,
        "current_user": current_user,
        "reagent_type": "mm",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("reagents/list.html", context)

@router.get("/mm/add", response_class=HTMLResponse)
async def add_mm_reagent_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add MM reagent form"""
    context = {
        "request": request,
        "title": "Add MM Reagent - EHS Electronic Journal",
        "current_user": current_user,
        "reagent_type": "mm"
    }
    
    return templates.TemplateResponse("reagents/add.html", context)

@router.post("/mm/api/", response_model=dict)
async def create_mm_reagent(
    reagent: MMReagentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new MM reagent"""
    
    # Check if batch number already exists
    existing = db.query(MMReagents).filter(MMReagents.batch_number == reagent.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        db_reagent = MMReagents(**reagent.dict(), prepared_by=current_user.id)
        db.add(db_reagent)
        db.commit()
        db.refresh(db_reagent)
        
        # Create history entry
        history_entry = MMReagentsHistory(
            reagent_id=db_reagent.id,
            action="created",
            new_value=f"MM Reagent {db_reagent.reagent_name} prepared",
            notes="Initial reagent preparation",
            remaining_volume=db_reagent.total_volume,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "MM Reagent created successfully",
            "reagent": db_reagent.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating MM reagent: {str(e)}"
        )

@router.get("/mm/export")
async def export_mm_reagents(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export MM reagents to Excel"""
    reagents = db.query(MMReagents).filter(MMReagents.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for reagent in reagents:
        data.append({
            'ID': reagent.id,
            'Reagent Name': reagent.reagent_name,
            'Batch Number': reagent.batch_number,
            'Preparation Date': reagent.preparation_date.strftime('%Y-%m-%d') if reagent.preparation_date else '',
            'Expiration Date': reagent.expiration_date.strftime('%Y-%m-%d') if reagent.expiration_date else '',
            'Total Volume (mL)': float(reagent.total_volume) if reagent.total_volume else 0,
            'Concentration': reagent.concentration or '',
            'pH Value': float(reagent.ph_value) if reagent.ph_value else '',
            'Conductivity': float(reagent.conductivity) if reagent.conductivity else '',
            'Prepared By': reagent.preparer.full_name if reagent.preparer else '',
            'Notes': reagent.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='MM Reagents', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['MM Reagents']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mm_reagents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

@router.get("/mm/api/", response_model=List[dict])
async def list_mm_reagents(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """List all MM reagents"""
    
    query = db.query(MMReagents)
    if active_only:
        query = query.filter(MMReagents.is_active == True)
    
    reagents = query.order_by(MMReagents.preparation_date.desc()).all()
    return [reagent.to_dict() for reagent in reagents]

@router.get("/mm/{reagent_id}", response_class=HTMLResponse)
async def mm_reagent_detail(
    reagent_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """MM reagent detail page"""
    
    reagent = db.query(MMReagents).filter(MMReagents.id == reagent_id).first()
    if not reagent:
        raise HTTPException(status_code=404, detail="Reagent not found")
    
    history = db.query(MMReagentsHistory).filter(
        MMReagentsHistory.reagent_id == reagent_id
    ).order_by(MMReagentsHistory.changed_at.desc()).all()
    
    context = {
        "request": request,
        "title": f"{reagent.reagent_name} - MM Reagent Details",
        "reagent": reagent,
        "history": history,
        "current_user": current_user,
        "reagent_type": "mm"
    }
    
    return templates.TemplateResponse("reagents/detail.html", context)

@router.put("/mm/api/{reagent_id}", response_model=dict)
async def update_mm_reagent(
    reagent_id: int,
    reagent_update: MMReagentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["update"]))
):
    """Update MM reagent"""
    
    db_reagent = db.query(MMReagents).filter(MMReagents.id == reagent_id).first()
    if not db_reagent:
        raise HTTPException(status_code=404, detail="Reagent not found")
    
    # Track changes for history
    changes = []
    update_data = reagent_update.dict(exclude_unset=True)
    
    for field, new_value in update_data.items():
        old_value = getattr(db_reagent, field)
        if old_value != new_value:
            changes.append({
                "field": field,
                "old_value": str(old_value) if old_value else None,
                "new_value": str(new_value) if new_value else None
            })
            setattr(db_reagent, field, new_value)
    
    if not changes:
        return {
            "success": True,
            "message": "No changes detected",
            "reagent": db_reagent.to_dict()
        }
    
    try:
        db.commit()
        
        # Create history entries for each change
        for change in changes:
            history_entry = MMReagentsHistory(
                reagent_id=reagent_id,
                action="updated",
                field_changed=change["field"],
                old_value=change["old_value"],
                new_value=change["new_value"],
                remaining_volume=db_reagent.total_volume,
                changed_by=current_user.id
            )
            db.add(history_entry)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"MM Reagent updated successfully. {len(changes)} field(s) modified.",
            "reagent": db_reagent.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error updating reagent: {str(e)}"
        )

@router.patch("/mm/api/{reagent_id}/volume", response_model=dict)
async def update_mm_volume(
    reagent_id: int,
    volume_update: VolumeUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["update"]))
):
    """Update MM reagent volume"""
    
    db_reagent = db.query(MMReagents).filter(MMReagents.id == reagent_id).first()
    if not db_reagent:
        raise HTTPException(status_code=404, detail="Reagent not found")
    
    old_volume = db_reagent.total_volume
    new_volume = old_volume + volume_update.volume_change
    
    if new_volume < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Insufficient volume available"
        )
    
    try:
        db_reagent.total_volume = new_volume
        db.commit()
        
        # Create history entry for volume change
        action = "volume_added" if volume_update.volume_change > 0 else "volume_used"
        history_entry = MMReagentsHistory(
            reagent_id=reagent_id,
            action=action,
            field_changed="total_volume",
            old_value=str(old_volume),
            new_value=str(new_volume),
            volume_used=volume_update.volume_change,
            remaining_volume=new_volume,
            reason=volume_update.reason,
            notes=volume_update.notes,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": f"Volume updated: {volume_update.volume_change:+.3f} mL",
            "reagent": db_reagent.to_dict(),
            "old_volume": float(old_volume),
            "new_volume": float(new_volume)
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error updating volume: {str(e)}"
        )

# Pb Reagents Routes (similar structure to MM)
@router.get("/pb", response_class=HTMLResponse)
async def pb_reagents_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Pb Reagents list page"""
    reagents = db.query(PbReagents).filter(
        PbReagents.is_active == True
    ).order_by(PbReagents.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "Pb Reagents - EHS Electronic Journal",
        "reagents": reagents,
        "current_user": current_user,
        "reagent_type": "pb",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("reagents/list.html", context)

@router.get("/pb/add", response_class=HTMLResponse)
async def add_pb_reagent_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add Pb reagent form"""
    context = {
        "request": request,
        "title": "Add Pb Reagent - EHS Electronic Journal",
        "current_user": current_user,
        "reagent_type": "pb"
    }
    
    return templates.TemplateResponse("reagents/add.html", context)

@router.post("/pb/api/", response_model=dict)
async def create_pb_reagent(
    reagent: PbReagentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new Pb reagent"""
    
    # Check if batch number already exists
    existing = db.query(PbReagents).filter(PbReagents.batch_number == reagent.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        db_reagent = PbReagents(**reagent.dict(), prepared_by=current_user.id)
        db.add(db_reagent)
        db.commit()
        db.refresh(db_reagent)
        
        # Create history entry
        history_entry = PbReagentsHistory(
            reagent_id=db_reagent.id,
            action="created",
            new_value=f"Pb Reagent {db_reagent.reagent_name} prepared",
            notes="Initial reagent preparation",
            remaining_volume=db_reagent.total_volume,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "Pb Reagent created successfully",
            "reagent": db_reagent.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating Pb reagent: {str(e)}"
        )

@router.get("/pb/export")
async def export_pb_reagents(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export Pb reagents to Excel"""
    reagents = db.query(PbReagents).filter(PbReagents.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for reagent in reagents:
        data.append({
            'ID': reagent.id,
            'Reagent Name': reagent.reagent_name,
            'Batch Number': reagent.batch_number,
            'Preparation Date': reagent.preparation_date.strftime('%Y-%m-%d') if reagent.preparation_date else '',
            'Expiration Date': reagent.expiration_date.strftime('%Y-%m-%d') if reagent.expiration_date else '',
            'Total Volume (mL)': float(reagent.total_volume) if reagent.total_volume else 0,
            'Concentration': reagent.concentration or '',
            'pH Value': float(reagent.ph_value) if reagent.ph_value else '',
            'Conductivity': float(reagent.conductivity) if reagent.conductivity else '',
            'Prepared By': reagent.preparer.full_name if reagent.preparer else '',
            'Notes': reagent.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Pb Reagents', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['Pb Reagents']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pb_reagents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# TCLP Reagents Routes (similar structure)
@router.get("/tclp", response_class=HTMLResponse)
async def tclp_reagents_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """TCLP Reagents list page"""
    reagents = db.query(TCLPReagents).filter(
        TCLPReagents.is_active == True
    ).order_by(TCLPReagents.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "TCLP Reagents - EHS Electronic Journal",
        "reagents": reagents,
        "current_user": current_user,
        "reagent_type": "tclp",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("reagents/list.html", context)

@router.get("/tclp/add", response_class=HTMLResponse)
async def add_tclp_reagent_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add TCLP reagent form"""
    context = {
        "request": request,
        "title": "Add TCLP Reagent - EHS Electronic Journal",
        "current_user": current_user,
        "reagent_type": "tclp"
    }
    
    return templates.TemplateResponse("reagents/add.html", context)

@router.post("/tclp/api/", response_model=dict)
async def create_tclp_reagent(
    reagent: TCLPReagentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new TCLP reagent"""
    
    # Check if batch number already exists
    existing = db.query(TCLPReagents).filter(TCLPReagents.batch_number == reagent.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        db_reagent = TCLPReagents(**reagent.dict(), prepared_by=current_user.id)
        db.add(db_reagent)
        db.commit()
        db.refresh(db_reagent)
        
        # Create history entry
        history_entry = TCLPReagentsHistory(
            reagent_id=db_reagent.id,
            action="created",
            new_value=f"TCLP Reagent {db_reagent.reagent_name} prepared",
            notes="Initial reagent preparation",
            remaining_volume=db_reagent.total_volume,
            changed_by=current_user.id
        )
        
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "TCLP Reagent created successfully",
            "reagent": db_reagent.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating TCLP reagent: {str(e)}"
        )

@router.get("/tclp/export")
async def export_tclp_reagents(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export TCLP reagents to Excel"""
    reagents = db.query(TCLPReagents).filter(TCLPReagents.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for reagent in reagents:
        data.append({
            'ID': reagent.id,
            'Reagent Name': reagent.reagent_name,
            'Batch Number': reagent.batch_number,
            'Reagent Type': reagent.reagent_type,
            'Preparation Date': reagent.preparation_date.strftime('%Y-%m-%d') if reagent.preparation_date else '',
            'Expiration Date': reagent.expiration_date.strftime('%Y-%m-%d') if reagent.expiration_date else '',
            'Total Volume (mL)': float(reagent.total_volume) if reagent.total_volume else 0,
            'pH Target': float(reagent.ph_target) if reagent.ph_target else '',
            'Final pH': float(reagent.final_ph) if reagent.final_ph else '',
            'Conductivity': float(reagent.conductivity) if reagent.conductivity else '',
            'Verification Passed': reagent.verification_passed,
            'Prepared By': reagent.preparer.full_name if reagent.preparer else '',
            'Notes': reagent.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='TCLP Reagents', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['TCLP Reagents']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tclp_reagents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# Generic routes for all reagent types
@router.get("/api/", response_model=dict)
async def list_all_reagents(
    reagent_type: Optional[str] = None,
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """List all reagents across types"""
    
    result = {
        "mm_reagents": [],
        "pb_reagents": [],
        "tclp_reagents": []
    }
    
    if not reagent_type or reagent_type.lower() == "mm":
        query = db.query(MMReagents)
        if active_only:
            query = query.filter(MMReagents.is_active == True)
        result["mm_reagents"] = [r.to_dict() for r in query.all()]
    
    if not reagent_type or reagent_type.lower() == "pb":
        query = db.query(PbReagents)
        if active_only:
            query = query.filter(PbReagents.is_active == True)
        result["pb_reagents"] = [r.to_dict() for r in query.all()]
    
    if not reagent_type or reagent_type.lower() == "tclp":
        query = db.query(TCLPReagents)
        if active_only:
            query = query.filter(TCLPReagents.is_active == True)
        result["tclp_reagents"] = [r.to_dict() for r in query.all()]
    
    if not reagent_type or reagent_type.lower() == "mercury":
        query = db.query(MercuryReagents)
        if active_only:
            query = query.filter(MercuryReagents.is_active == True)
        result["mercury_reagents"] = [r.to_dict() for r in query.all()]
    
    return result

# Mercury Reagents Routes
@router.get("/mercury", response_class=HTMLResponse)
async def mercury_reagents_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Mercury Reagents list page"""
    reagents = db.query(MercuryReagents).filter(
        MercuryReagents.is_active == True
    ).order_by(MercuryReagents.preparation_date.desc()).all()
    
    context = {
        "request": request,
        "title": "Mercury Reagents - EHS Electronic Journal",
        "reagents": reagents,
        "current_user": current_user,
        "reagent_type": "Mercury",
        "today": datetime.now().date()
    }
    
    return templates.TemplateResponse("reagents/list.html", context)

@router.get("/mercury/add", response_class=HTMLResponse)
async def add_mercury_reagent_form(
    request: Request,
    current_user: User = Depends(require_permissions(["create"]))
):
    """Add Mercury reagent form"""
    context = {
        "request": request,
        "title": "Add Mercury Reagent - EHS Electronic Journal",
        "current_user": current_user,
        "reagent_type": "Mercury"
    }
    
    return templates.TemplateResponse("reagents/add.html", context)

@router.post("/mercury/api/", response_model=dict)
async def create_mercury_reagent(
    reagent: MercuryReagentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["create"]))
):
    """Create new Mercury reagent"""
    
    # Check if batch number already exists
    existing = db.query(MercuryReagents).filter(MercuryReagents.batch_number == reagent.batch_number).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Batch number already exists"
        )
    
    try:
        db_reagent = MercuryReagents(**reagent.dict(), prepared_by=current_user.id)
        db.add(db_reagent)
        db.commit()
        db.refresh(db_reagent)
        
        # Create history entry
        history_entry = MercuryReagentsHistory(
            reagent_id=db_reagent.id,
            action="created",
            new_value=f"Mercury Reagent {db_reagent.reagent_name} prepared",
            notes="Initial reagent preparation",
            changed_by=current_user.id
        )
        db.add(history_entry)
        db.commit()
        
        return {
            "success": True,
            "message": "Mercury Reagent created successfully",
            "reagent": db_reagent.to_dict()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error creating Mercury reagent: {str(e)}"
        )

@router.get("/mercury/export")
async def export_mercury_reagents(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permissions(["read"]))
):
    """Export Mercury reagents to Excel"""
    reagents = db.query(MercuryReagents).filter(MercuryReagents.is_active == True).all()
    
    # Convert to DataFrame
    data = []
    for reagent in reagents:
        data.append({
            'ID': reagent.id,
            'Reagent Name': reagent.reagent_name,
            'Batch Number': reagent.batch_number,
            'Preparation Date': reagent.preparation_date.strftime('%Y-%m-%d') if reagent.preparation_date else '',
            'Expiration Date': reagent.expiration_date.strftime('%Y-%m-%d') if reagent.expiration_date else '',
            'Total Volume (mL)': float(reagent.total_volume) if reagent.total_volume else 0,
            'Concentration': reagent.concentration or '',
            'pH Value': float(reagent.ph_value) if reagent.ph_value else '',
            'Conductivity': float(reagent.conductivity) if reagent.conductivity else '',
            'Prepared By': reagent.preparer.full_name if reagent.preparer else '',
            'Notes': reagent.notes or ''
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Mercury Reagents', index=False)
        
        # Style the worksheet
        worksheet = writer.sheets['Mercury Reagents']
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mercury_reagents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )